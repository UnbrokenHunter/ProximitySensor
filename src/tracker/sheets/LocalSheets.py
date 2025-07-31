import os
import time
import openpyxl
from openpyxl import Workbook

from .. import Globals
from .. import Statistics

import threading

excel_lock = threading.Lock()

SPREADSHEET_FILE_PATH = "runtime/sheets_backup.xlsx"
SHEET_NAME = "Sheet1"

def get_or_create_workbook(file_path):
    """Check if the spreadsheet file exists. If not, create it and add headers."""
    headers = ["Lap Count", "Lap Time", "Driver Name", "Time"]
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(headers)  # Add headers to the first row
        wb.save(file_path)
    else:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[SHEET_NAME]
        if sheet.max_row == 1:
            # Check if the first row is empty or only contains headers; if so, append the headers.
            empty_or_headers_only = all(sheet[f"A1":f"G1"][0][i].value in (None, headers[i]) for i in range(len(headers)))
            if empty_or_headers_only:
                for col_num, header in enumerate(headers, start=1):
                    sheet.cell(row=1, column=col_num, value=header)
                wb.save(file_path)
    return wb

def update_cell(row_number, column_letter, new_value):
    """Updates a specific cell in a spreadsheet."""
    
    with excel_lock:
        wb = get_or_create_workbook(SPREADSHEET_FILE_PATH)
        sheet = wb[SHEET_NAME]
        cell_reference = f"{column_letter}{row_number + 1}"
        sheet[cell_reference] = new_value
        wb.save(SPREADSHEET_FILE_PATH)
    
    if Globals.EnableLogging:
        print(f"Updated local cell {cell_reference}.")

def find_first_empty_cell_in_column(column_letter='A'):
    """Finds the first empty cell in a specified column of a local Excel sheet."""
    
    with excel_lock:
        wb = get_or_create_workbook(SPREADSHEET_FILE_PATH)
        sheet = wb[SHEET_NAME]
        column = column_letter.upper()  # Ensure the column letter is uppercase for consistency
        
        row_number = 0
        for row in sheet[column]:
            if row.value is not None:
                row_number += 1
            else:
                break
        
    if Globals.EnableLogging:
        # Log or handle the first empty cell's position as needed. For debugging, you might want to print it.
        print(f"First empty cell in column {column_letter} is at row {row_number}")
    
    return row_number

def SaveDataManual(Min, LapCount, LapTime, Driver, Time):
    try:
        minRow = Min

        # Lap Count
        update_cell(minRow, 'A', LapCount)

        # Lap Time
        update_cell(minRow, 'B', LapTime)

        # Driver Name
        update_cell(minRow, 'C', Driver)

        # Time
        update_cell(minRow, 'D', Time)

    except Exception as err:
        print(err)

def SaveData(Min, LapTime):
    SaveDataManual(Min, Min, LapTime, Globals.CurrentDriver, time.strftime("%Y-%m-%d %H:%M:%S"))

def read_cell(row_number, column_letter):
    """Reads a specific cell from the local spreadsheet."""
        
    with excel_lock:
        wb = get_or_create_workbook(SPREADSHEET_FILE_PATH)
        sheet = wb[SHEET_NAME]
        cell_reference = f"{column_letter.upper()}{row_number}"
        value = sheet[cell_reference].value
    
    if Globals.EnableLogging:
        print(f"Read cell {cell_reference}: {value}")
        
    return value

def get_last_n_laps(n=5):
    with excel_lock:
        wb = get_or_create_workbook(SPREADSHEET_FILE_PATH)
        sheet = wb[SHEET_NAME]
        data = []

        max_row = sheet.max_row
        headers = [cell.value for cell in sheet[1]]
        
        # Read last `n` rows excluding the header row
        for i in range(max(2, max_row - n + 1), max_row + 1):
            row = {
                "lap_count": sheet[f"A{i}"].value,
                "lap_time": sheet[f"B{i}"].value,
                "driver": sheet[f"C{i}"].value,
                "timestamp": sheet[f"D{i}"].value,
            }
            data.append(row)

    return data[::-1]  # Return in newest-to-oldest order
