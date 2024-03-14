import os
import openpyxl
from openpyxl import Workbook
import time

import Globals
import Statistics

SPREADSHEET_FILE_PATH = "sheets_backup.xlsx"

def get_or_create_workbook(file_path):
    """Check if the spreadsheet file exists. If not, create it and add headers."""
    headers = ["Lap Count", "Lap Time", "Driver Name", "Distance Driven", "Instant Speed", "Average Speed", "Time"]
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(headers)  # Add headers to the first row
        wb.save(file_path)
    else:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb["Sheet1"]
        if sheet.max_row == 1:
            # Check if the first row is empty or only contains headers; if so, append the headers.
            empty_or_headers_only = all(sheet[f"A1":f"G1"][0][i].value in (None, headers[i]) for i in range(len(headers)))
            if empty_or_headers_only:
                for col_num, header in enumerate(headers, start=1):
                    sheet.cell(row=1, column=col_num, value=header)
                wb.save(file_path)
    return wb

def update_cell(sheet_name, row_number, column_letter, new_value):
    """Updates a specific cell in a spreadsheet."""
    wb = get_or_create_workbook(SPREADSHEET_FILE_PATH)
    sheet = wb[sheet_name]
    cell_reference = f"{column_letter}{row_number + 1}"
    sheet[cell_reference] = new_value
    wb.save(SPREADSHEET_FILE_PATH)
    print(f"Updated local cell {cell_reference}.")

def find_first_empty_cell_in_column(sheet_name, column_letter='A'):
    """Finds the first empty cell in a specified column of a local Excel sheet."""
    wb = get_or_create_workbook(SPREADSHEET_FILE_PATH)
    sheet = wb[sheet_name]
    column = column_letter.upper()  # Ensure the column letter is uppercase for consistency
    
    row_number = 1
    for row in sheet[column]:
        if row.value is not None:
            row_number += 1
        else:
            break
    
    # Adjust Globals.LapCount if necessary. Ensure Globals.LapCount is properly initialized and managed in your script.
    if Globals.ControlsLapCount == "Local":
        Globals.LapCount = row_number - 1
    
    # Log or handle the first empty cell's position as needed. For debugging, you might want to print it.
    print(f"First empty cell in column {column_letter} is at row {row_number}")
    
    return row_number

def SaveDataManual(LapTime, Driver, DistanceDriven, InstantSpeed, Time):
    try:
        sheet_name = "Sheet1"
        minRow = find_first_empty_cell_in_column(sheet_name)

        # Lap Count
        update_cell(sheet_name, minRow, 'A', Globals.LapCount)

        # Lap Time
        update_cell(sheet_name, minRow, 'B', LapTime)

        # Driver Name
        update_cell(sheet_name, minRow, 'C', Driver)

        # Distance Driven
        distance_driven = f"{DistanceDriven}km"
        update_cell(sheet_name, minRow, 'D', distance_driven)

        # Instant Speed
        instant_speed_mph = f"{((Globals.CarLength / InstantSpeed) * 3600) * 0.621371192}"  # From Car Length / Time To Drive that Distance to Km / H to Mph
        update_cell(sheet_name, minRow, 'E', instant_speed_mph)

        # Average Speed
        average_speed_mph = f"{((DistanceDriven / LapTime) * 3600) * 0.621371192}"  # From Km / S to Km / H to Mph
        update_cell(sheet_name, minRow, 'F', average_speed_mph)

        # Time
        update_cell(sheet_name, minRow, 'G', Time)

    except Exception as err:
        print(err)

def SaveData(LapTime, InstantSpeed):
    SaveDataManual(LapTime, Globals.CurrentDriver, Statistics.GetDistanceDriven(), InstantSpeed, time.strftime("%Y-%m-%d %H:%M:%S"))

